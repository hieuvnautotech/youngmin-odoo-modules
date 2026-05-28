import json
from odoo import http
from odoo.http import request
import xlsxwriter
from datetime import datetime
import os
import openpyxl
from io import BytesIO

from odoo.exceptions import UserError
from odoo import _


class QMSQC_Print(http.Controller):

    @http.route("/pms/print/iqc",type="http",auth="user")
    def action_print_iqc(self, state=None, **kwargs):
        if not state:
            raise UserError(_("Missing state"))
        data = json.loads(state)
        module_path = os.path.dirname(os.path.dirname(__file__))
        template_path = os.path.join(module_path, "report", "IQC.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(_(f"Template file not found: {template_path}"))
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.worksheets[0]

        sheet[f"A1"] = data.get("title") or ""

        sheet[f"B2"] = data.get("category_material") or ""
        sheet[f"B3"] = data.get("supplier") or ""
        sheet[f"E2"] = data.get("material_code") or ""
        sheet[f"E3"] = data.get("material_name") or ""
        sheet[f"H2"] = data.get("lot") or ""
        sheet[f"H3"] = f"{data.get('qty', '')} {data.get('qty_uom', '')}"

        sheet[f"E4"] = "Staff:" + data.get("staff_name") or ""
        sheet[f"H4"] = "Date:" + data.get("check_date") or ""
        sheet[f"K4"] = "Final Result:" + data.get("final_result") or ""

        row = 6
        columns = data.get('columns', [])
        for i, col in enumerate(columns[:6]):
            sheet.cell(row=row, column=i + 1).value = col.get('title', '')
        for line in data.get('check_list', []):
            row += 1
            sheet[f"A{row}"] = line.get('qc_code', '')
            sheet[f"B{row}"] = line.get('method', '')
            sheet[f"C{row}"] = line.get('frequency', '')
            sheet[f"D{row}"] = line.get('standard', '')
            sheet[f"E{row}"] = line.get('X1', '')
            sheet[f"F{row}"] = line.get('X2', '')
            sheet[f"G{row}"] = line.get('X3', '')
            sheet[f"H{row}"] = line.get('X4', '')
            sheet[f"I{row}"] = line.get('X5', '')
            sheet[f"J{row}"] = line.get('result', '')
            sheet[f"K{row}"] = line.get('remark', '')

        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"IQC_{current_datetime}.xlsx"
        return request.make_response(
            file_data.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                (f'Content-Disposition', f'attachment; filename={filename};')
            ]
        )

    @http.route("/pms/print/item_qc", type="http", auth="user")
    def action_print_item_qc(self, state=None, **kwargs):
        if not state:
            raise UserError(_("Missing state"))
        data = json.loads(state)
        module_path = os.path.dirname(os.path.dirname(__file__))
        template_path = os.path.join(module_path, "report", "Item_QC.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(_(f"Template file not found: {template_path}"))
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.worksheets[0]

        sheet[f"A1"] = data.get("title") or ""
        sheet[f"B2"] = data.get("material_code") or ""
        sheet[f"E2"] = data.get("material_name") or ""
        sheet[f"H2"] = data.get("lot") or ""
        sheet[f"K2"] = data.get("spec") or ""

        sheet[f"H3"] = "Staff:" + data.get("staff_name") or ""
        sheet[f"K3"] = "Date:" + data.get("check_date") or ""
        sheet[f"N3"] = "Final Result:" + data.get("final_result") or ""

        row = 5
        columns = data.get('columns', [])
        for i, col in enumerate(columns[:6]):
            sheet.cell(row=row, column=i + 1).value = col.get('title', '')
        for line in data.get('check_list', []):
            row += 1
            sheet[f"A{row}"] = line.get('qc_code', '')
            sheet[f"B{row}"] = line.get('method', '')
            sheet[f"C{row}"] = line.get('frequency', '')
            sheet[f"D{row}"] = line.get('standard', '')
            sheet[f"E{row}"] = line.get('result', '')
            sheet[f"F{row}"] = line.get('remark', '')

        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Item_QC_{current_datetime}.xlsx"
        return request.make_response(
            file_data.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                (f'Content-Disposition', f'attachment; filename={filename};')
            ]
        )

    @http.route("/pms/print/jig", type="http", auth="user")
    def action_print_jig(self, state=None, **kwargs):
        if not state:
            raise UserError(_("Missing state"))
        data = json.loads(state)
        module_path = os.path.dirname(os.path.dirname(__file__))
        template_path = os.path.join(module_path, "report", "Jig.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(_(f"Template file not found: {template_path}"))
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.worksheets[0]

        sheet[f"A1"] = data.get("title") or ""

        sheet[f"B2"] = data.get("jig_code") or ""
        sheet[f"B3"] = data.get("jig_name") or ""
        sheet[f"E2"] = data.get("product_code") or ""
        sheet[f"E3"] = data.get("supplier") or ""
        sheet[f"H2"] = data.get("process") or ""
        sheet[f"H3"] = data.get("rev") or ""
        sheet[f"K2"] = data.get("remark") or ""
        sheet[f"K3"] = data.get("create_date") or ""

        sheet[f"E4"] = "Staff:" + data.get("staff_name") or ""
        sheet[f"H4"] = "Date:" + data.get("check_date") or ""
        sheet[f"K4"] = "Final Result:" + data.get("final_result") or ""

        row = 5
        columns = data.get('columns', [])
        for i, col in enumerate(columns[:6]):
            sheet.cell(row=row, column=i + 1).value = col.get('title', '')
        for line in data.get('check_list', []):
            row += 1
            sheet[f"A{row}"] = line.get('qc_code', '')
            sheet[f"B{row}"] = line.get('method', '')
            sheet[f"C{row}"] = line.get('frequency', '')
            sheet[f"D{row}"] = line.get('standard', '')
            sheet[f"E{row}"] = line.get('result', '')
            sheet[f"F{row}"] = line.get('remark', '')

        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Jig_{current_datetime}.xlsx"
        return request.make_response(
            file_data.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                (f'Content-Disposition', f'attachment; filename={filename};')
            ]
        )

    @http.route("/pms/print/oqc", type="http", auth="user")
    def action_print_oqc(self, state=None, **kwargs):
        if not state:
            raise UserError(_("Missing state"))
        data = json.loads(state)
        module_path = os.path.dirname(os.path.dirname(__file__))
        template_path = os.path.join(module_path, "report", "OQC.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(_(f"Template file not found: {template_path}"))
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.worksheets[0]

        sheet[f"A1"] = data.get("title") or ""

        sheet[f"B2"] = data.get("material_code") or ""
        sheet[f"B3"] = data.get("lot") or ""
        sheet[f"E2"] = data.get("material_name") or ""
        sheet[f"E3"] = data.get("line_no") or ""
        sheet[f"H2"] = data.get("qty_sampling") or ""
        sheet[f"H3"] = data.get("ng_qty") or ""
        sheet[f"K2"] = data.get("defect_ratio") or ""
        sheet[f"K3"] = data.get("ok_qty") or ""

        sheet[f"A4"] = "QC Three-Stage inspection: " + data.get("type_roll") or ""
        sheet[f"E4"] = "Staff: " + data.get("staff_name") or ""
        sheet[f"H4"] = "Date: " + data.get("check_date") or ""

        row = 5
        columns = data.get('columns', [])
        for i, col in enumerate(columns[:6]):
            sheet.cell(row=row, column=i + 1).value = col.get('title', '')
        for line in data.get('check_list', []):
            row += 1
            sheet[f"A{row}"] = line.get('qc_code', '')
            sheet[f"B{row}"] = line.get('method', '')
            sheet[f"C{row}"] = line.get('frequency', '')
            sheet[f"D{row}"] = line.get('standard', '')
            sheet[f"E{row}"] = line.get('result', '')
            sheet[f"F{row}"] = line.get('remark', '')

        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"OQC_{current_datetime}.xlsx"
        return request.make_response(
            file_data.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                (f'Content-Disposition', f'attachment; filename={filename};')
            ]
        )

    @http.route("/pms/print/pqc", type="http", auth="user")
    def action_print_pqc(self, state=None, **kwargs):
        if not state:
            raise UserError(_("Missing state"))
        data = json.loads(state)
        module_path = os.path.dirname(os.path.dirname(__file__))
        template_path = os.path.join(module_path, "report", "OQC.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(_(f"Template file not found: {template_path}"))
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.worksheets[0]

        sheet[f"A1"] = data.get("title") or ""

        sheet[f"B2"] = data.get("material_code") or ""
        sheet[f"B3"] = data.get("lot") or ""
        sheet[f"E2"] = data.get("material_name") or ""
        sheet[f"E3"] = data.get("line_no") or ""
        sheet[f"H2"] = data.get("qty_sampling") or ""
        sheet[f"H3"] = data.get("ng_qty") or ""
        sheet[f"K2"] = data.get("defect_ratio") or ""
        sheet[f"K3"] = data.get("ok_qty") or ""

        sheet[f"A4"] = "QC Three-Stage inspection: " + data.get("type_roll") or ""
        sheet[f"E4"] = "Staff: " + data.get("staff_name") or ""
        sheet[f"H4"] = "Date: " + data.get("check_date") or ""

        row = 5
        columns = data.get('columns', [])
        for i, col in enumerate(columns):
            sheet.cell(row=row, column=i + 1).value = col.get('title', '')
        if len(columns) > 10 :
            for line in data.get('check_list', []):
                row += 1
                sheet[f"A{row}"] = line.get('qc_code', '')
                sheet[f"B{row}"] = line.get('method', '')
                sheet[f"C{row}"] = line.get('frequency', '')
                sheet[f"D{row}"] = line.get('standard', '')
                sheet[f"E{row}"] = line.get('HD1', '')
                sheet[f"F{row}"] = line.get('HD2', '')
                sheet[f"G{row}"] = line.get('HD3', '')
                sheet[f"H{row}"] = line.get('HD4', '')
                sheet[f"I{row}"] = line.get('HD5', '')
                sheet[f"J{row}"] = line.get('HD6', '')
                sheet[f"K{row}"] = line.get('HD7', '')
                sheet[f"L{row}"] = line.get('HD8', '')
                sheet[f"M{row}"] = line.get('HD9', '')
                sheet[f"N{row}"] = line.get('HD10', '')
                sheet[f"O{row}"] = line.get('HD11', '')
                sheet[f"P{row}"] = line.get('HD12', '')
                sheet[f"Q{row}"] = line.get('HD13', '')
                sheet[f"R{row}"] = line.get('HD14', '')
                sheet[f"S{row}"] = line.get('remark', '')
        else:
            for line in data.get('check_list', []):
                row += 1
                sheet[f"A{row}"] = line.get('qc_code', '')
                sheet[f"B{row}"] = line.get('method', '')
                sheet[f"C{row}"] = line.get('frequency', '')
                sheet[f"D{row}"] = line.get('standard', '')
                sheet[f"E{row}"] = line.get('result', '')
                sheet[f"F{row}"] = line.get('remark', '')

        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"QC_{current_datetime}.xlsx"
        return request.make_response(
            file_data.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                (f'Content-Disposition', f'attachment; filename={filename};')
            ]
        )

    @http.route("/pms/print/pallet_qc", type="http", auth="user")
    def action_print_pallet_qc(self, state=None, **kwargs):
        if not state:
            raise UserError(_("Missing state"))
        data = json.loads(state)
        module_path = os.path.dirname(os.path.dirname(__file__))
        template_path = os.path.join(module_path, "report", "Pallet_QC.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(_(f"Template file not found: {template_path}"))
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.worksheets[0]

        sheet[f"A1"] = data.get("title") or ""

        sheet[f"B2"] = data.get("pallet_no") or ""
        sheet[f"B3"] = data.get("project_code") or ""
        sheet[f"E2"] = data.get("material_code") or ""
        sheet[f"E3"] = data.get("material_name") or ""
        sheet[f"H2"] = data.get("lot") or ""
        sheet[f"H3"] = f"{data.get('qty', '')} {data.get('qty_uom', '')}"

        sheet[f"E4"] = "Staff:" + data.get("staff_name") or ""
        sheet[f"H4"] = "Date:" + data.get("check_date") or ""
        sheet[f"K4"] = "Final Result:" + data.get("final_result") or ""

        row = 6
        columns = data.get('columns', [])
        for i, col in enumerate(columns[:6]):
            sheet.cell(row=row, column=i + 1).value = col.get('title', '')
        for line in data.get('check_list', []):
            row += 1
            sheet[f"A{row}"] = line.get('qc_code', '')
            sheet[f"B{row}"] = line.get('method', '')
            sheet[f"C{row}"] = line.get('frequency', '')
            sheet[f"D{row}"] = line.get('standard', '')
            sheet[f"E{row}"] = line.get('result', '')
            sheet[f"F{row}"] = line.get('remark', '')

        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Pallet_QC_{current_datetime}.xlsx"
        return request.make_response(
            file_data.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                (f'Content-Disposition', f'attachment; filename={filename};')
            ]
        )

    @http.route("/pms/print/re_use", type="http", auth="user")
    def action_print_re_use(self, state=None, **kwargs):
        if not state:
            raise UserError(_("Missing state"))
        data = json.loads(state)
        module_path = os.path.dirname(os.path.dirname(__file__))
        template_path = os.path.join(module_path, "report", "Re_use.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(_(f"Template file not found: {template_path}"))
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.worksheets[0]

        sheet[f"A1"] = data.get("title") or ""

        sheet[f"B2"] = data.get("material_code") or ""
        sheet[f"B3"] = data.get("lot") or ""
        sheet[f"E2"] = data.get("material_name") or ""
        sheet[f"E3"] = f"{data.get('qty', '')} {data.get('qty_uom', '')}"
        sheet[f"H2"] = data.get("remaining_days") or ""
        sheet[f"H3"] = data.get("life_days") or ""
        if data.get("item_group") == 'material':
            sheet[f"A2"] = "Material Code"
            sheet[f"D2"] = "Material Name"
        elif data.get("item_group") == 'semi':
            sheet[f"A2"] = "Semi Code"
            sheet[f"D2"] = "Semi Name"
        elif data.get("item_group") == 'fg':
            sheet[f"A2"] = "FG Code"
            sheet[f"D2"] = "FG Name"

        sheet[f"E4"] = "Staff:" + data.get("staff_name") or ""
        sheet[f"H4"] = "Date:" + data.get("check_date") or ""
        sheet[f"K4"] = "Final Result:" + data.get("final_result") or ""

        row = 6
        columns = data.get('columns', [])
        for i, col in enumerate(columns[:6]):
            sheet.cell(row=row, column=i + 1).value = col.get('title', '')
        for line in data.get('check_list', []):
            row += 1
            sheet[f"A{row}"] = line.get('qc_code', '')
            sheet[f"B{row}"] = line.get('method', '')
            sheet[f"C{row}"] = line.get('frequency', '')
            sheet[f"D{row}"] = line.get('standard', '')
            sheet[f"E{row}"] = line.get('result', '')
            sheet[f"F{row}"] = line.get('remark', '')

        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Re_use_{current_datetime}.xlsx"
        return request.make_response(
            file_data.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                (f'Content-Disposition', f'attachment; filename={filename};')
            ]
        )
